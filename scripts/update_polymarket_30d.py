#!/usr/bin/env python3
"""Update the Polymarket 30-day topic volume dashboard data."""

from __future__ import annotations

import argparse
import concurrent.futures
import csv
import json
import time
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "docs" / "data"
EVENTS_URL = "https://gamma-api.polymarket.com/events"
KALSHI_MARKETS_URL = "https://api.elections.kalshi.com/trade-api/v2/markets"
KALSHI_EVENTS_URL = "https://api.elections.kalshi.com/trade-api/v2/events"
HEADERS = {"User-Agent": "Mozilla/5.0", "Accept": "application/json"}

TOPICS = {
    "Sports": [
        "sports",
        "soccer",
        "basketball",
        "football",
        "baseball",
        "tennis",
        "ufc",
        "mma",
        "fifa-world-cup",
        "nba",
        "nfl",
        "mlb",
        "nhl",
        "olympics",
        "epl",
        "esports",
        "games",
    ],
    "Crypto": [
        "crypto",
        "cryptocurrency",
        "bitcoin",
        "ethereum",
        "solana",
        "xrp",
        "dogecoin",
        "memecoins",
        "airdrops",
    ],
    "Business/Economy": [
        "business",
        "economy",
        "finance",
        "fed",
        "fomc",
        "fed-rates",
        "markets",
        "companies",
        "stocks",
        "ipo",
        "earnings-calls",
    ],
    "Geopolitics": [
        "ukraine-russia",
        "ukraine-&-russia",
        "ukraine",
        "russia",
        "israel",
        "iran",
        "middle-east",
        "china",
        "geopolitics",
        "war",
        "israel-strike-iran",
    ],
    "Politics": [
        "politics",
        "us-current-affairs",
        "elections",
        "election",
        "trump",
        "biden",
        "congress",
        "senate",
        "president",
        "nyc-mayor",
        "macro-election-2",
    ],
    "Culture": [
        "pop-culture",
        "culture",
        "entertainment",
        "music",
        "movies",
        "television",
        "celebrity",
        "celebrities",
        "oscars",
        "tweet-markets",
        "tweets-markets",
    ],
    "Tech & Science": [
        "technology",
        "tech",
        "science",
        "ai",
        "artificial-intelligence",
        "openai",
        "spacex",
        "gpt-5pt5",
    ],
    "Health": ["health", "coronavirus", "covid-19"],
}

TOPIC_ORDER = [
    "Sports",
    "Politics",
    "Crypto",
    "Business/Economy",
    "Geopolitics",
    "Culture",
    "Tech & Science",
    "Health",
    "Unclassified",
]

CLASSIFY_ORDER = [
    "Sports",
    "Crypto",
    "Business/Economy",
    "Geopolitics",
    "Politics",
    "Culture",
    "Tech & Science",
    "Health",
]

IGNORE_TAGS = {
    "hide-from-new",
    "recurring",
    "breaking-news",
    "new",
    "featured",
    "homepage",
    "carousel",
    "daily",
    "weekly",
    "monthly",
    "hourly",
    "parent-for-derivative",
    "hfc",
}

KALSHI_CATEGORY_TOPICS = {
    "elections": "Politics",
    "politics": "Politics",
    "sports": "Sports",
    "crypto": "Crypto",
    "cryptocurrency": "Crypto",
    "economics": "Business/Economy",
    "financials": "Business/Economy",
    "companies": "Business/Economy",
    "business": "Business/Economy",
    "world": "Geopolitics",
    "climate and weather": "Health",
    "weather": "Health",
    "entertainment": "Culture",
    "science and technology": "Tech & Science",
    "technology": "Tech & Science",
}


def fetch_events(limit: int, max_rows: int) -> list[dict]:
    events: list[dict] = []
    for offset in range(0, max_rows, limit):
        params = urllib.parse.urlencode(
            {
                "limit": limit,
                "offset": offset,
                "order": "volume1mo",
                "ascending": "false",
            }
        )
        url = f"{EVENTS_URL}?{params}"
        for attempt in range(4):
            try:
                req = urllib.request.Request(url, headers=HEADERS)
                with urllib.request.urlopen(req, timeout=90) as response:
                    batch = json.load(response)
                break
            except Exception:
                if attempt == 3:
                    raise
                time.sleep(1.5 * (attempt + 1))
        events.extend(batch)
        if len(batch) < limit:
            break
        time.sleep(0.02)
    return events


def fetch_json(url: str, timeout: int = 90) -> dict:
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=timeout) as response:
        return json.load(response)


def classify_event(event: dict) -> str:
    tags = event.get("tags") or []
    slugs = [
        (tag.get("slug") or "").lower()
        for tag in tags
        if (tag.get("slug") or "").lower() not in IGNORE_TAGS
    ]
    labels = [(tag.get("label") or "").lower() for tag in tags]
    category = (event.get("category") or "").strip().lower().replace(" ", "-")
    haystack = set(slugs + labels + ([category] if category else []))

    for topic in CLASSIFY_ORDER:
        if any(alias in haystack for alias in TOPICS[topic]):
            return topic

    title = (event.get("title") or "").lower()
    if any(token in title for token in ["bitcoin", "ethereum", "crypto", "solana", "xrp", "dogecoin"]):
        return "Crypto"
    if any(token in title for token in ["fed decision", "interest rate", "fomc", "ipo", "earnings"]):
        return "Business/Economy"
    if any(token in title for token in ["iran", "israel", "russia", "ukraine", "china strikes", "war"]):
        return "Geopolitics"
    if any(token in title for token in ["trump", "biden", "election", "senate", "president", "congress", "mayor"]):
        return "Politics"
    if any(token in title for token in ["nba", "nfl", "world cup", "champion", "ufc", "soccer", "mlb", "nhl", "vs."]):
        return "Sports"
    return "Unclassified"


def summarize(events: list[dict], as_of: str) -> dict:
    volume_by_topic: dict[str, float] = defaultdict(float)
    count_by_topic: dict[str, int] = defaultdict(int)
    examples: dict[str, dict] = {}

    for event in events:
        volume = float(event.get("volume1mo") or 0)
        topic = classify_event(event)
        volume_by_topic[topic] += volume
        count_by_topic[topic] += 1
        examples.setdefault(
            topic,
            {
                "title": event.get("title"),
                "volume1mo": volume,
                "tags": [
                    [tag.get("label"), tag.get("slug")]
                    for tag in (event.get("tags") or [])[:8]
                ],
            },
        )

    total = sum(volume_by_topic.values())
    rows = []
    for topic in sorted(volume_by_topic, key=volume_by_topic.get, reverse=True):
        volume = volume_by_topic[topic]
        rows.append(
            {
                "date": as_of,
                "topic": topic,
                "volume1mo": volume,
                "share_pct": (volume / total * 100) if total else 0,
                "event_count": count_by_topic[topic],
                "example": examples.get(topic),
            }
        )

    return {
        "as_of": as_of,
        "platform": "Polymarket",
        "metric_label": "近 30 天成交量",
        "source": EVENTS_URL,
        "params": {
            "order": "volume1mo",
            "ascending": "false",
            "limit": 20,
            "offset": "0..1980",
        },
        "rows_fetched": len(events),
        "sum_volume1mo": total,
        "last_event_volume1mo": float(events[-1].get("volume1mo") or 0) if events else 0,
        "topics": rows,
    }


def fetch_kalshi_markets(limit: int, max_pages: int) -> list[dict]:
    markets: list[dict] = []
    cursor = ""
    for page in range(max_pages):
        params = {"limit": limit}
        if cursor:
            params["cursor"] = cursor
        url = f"{KALSHI_MARKETS_URL}?{urllib.parse.urlencode(params)}"
        for attempt in range(4):
            try:
                data = fetch_json(url, timeout=90)
                break
            except Exception:
                if attempt == 3:
                    raise
                time.sleep(1.5 * (attempt + 1))
        batch = data.get("markets") or []
        markets.extend(batch)
        cursor = data.get("cursor") or ""
        if not cursor or len(batch) < limit:
            break
        time.sleep(0.02)
    return markets


def fetch_kalshi_event_categories(event_tickers: set[str]) -> dict[str, str]:
    def fetch_category(ticker: str) -> tuple[str, str]:
        if not ticker:
            return ticker, ""
        url = f"{KALSHI_EVENTS_URL}/{urllib.parse.quote(ticker)}"
        for attempt in range(3):
            try:
                data = fetch_json(url, timeout=45)
                event = data.get("event") or {}
                return ticker, event.get("category") or ""
            except Exception:
                if attempt == 2:
                    return ticker, ""
                else:
                    time.sleep(0.8 * (attempt + 1))
        return ticker, ""

    categories: dict[str, str] = {}
    tickers = sorted(ticker for ticker in event_tickers if ticker)
    with concurrent.futures.ThreadPoolExecutor(max_workers=16) as executor:
        for ticker, category in executor.map(fetch_category, tickers):
            categories[ticker] = category
    return categories


def classify_kalshi_market(market: dict, event_category: str) -> str:
    category = (event_category or "").strip().lower()
    if category in KALSHI_CATEGORY_TOPICS:
        return KALSHI_CATEGORY_TOPICS[category]

    ticker = f"{market.get('event_ticker') or ''} {market.get('ticker') or ''}".lower()
    title = (market.get("title") or "").lower()
    haystack = f"{ticker} {title}"
    if any(token in haystack for token in ["nba", "nfl", "mlb", "nhl", "wnba", "atp", "ufc", "sports", "soccer", "tennis", "fifa"]):
        return "Sports"
    if any(token in haystack for token in ["bitcoin", "ethereum", "crypto", "solana", "xrp", "doge"]):
        return "Crypto"
    if any(token in haystack for token in ["fed", "cpi", "inflation", "rate", "gdp", "earnings", "ipo", "stock"]):
        return "Business/Economy"
    if any(token in haystack for token in ["election", "trump", "biden", "senate", "congress", "president", "governor"]):
        return "Politics"
    if any(token in haystack for token in ["iran", "israel", "ukraine", "russia", "china", "nato", "war"]):
        return "Geopolitics"
    if any(token in haystack for token in ["movie", "music", "oscar", "grammy", "celebrity"]):
        return "Culture"
    if any(token in haystack for token in ["ai", "openai", "spacex", "tesla", "technology"]):
        return "Tech & Science"
    return "Unclassified"


def summarize_kalshi(markets: list[dict], as_of: str, top_rows: int) -> dict:
    positive_markets = [
        market for market in markets if float(market.get("volume_fp") or 0) > 0
    ]
    positive_markets.sort(key=lambda market: float(market.get("volume_fp") or 0), reverse=True)
    selected = positive_markets[:top_rows]
    event_categories = fetch_kalshi_event_categories(
        {market.get("event_ticker") or "" for market in selected}
    )

    volume_by_topic: dict[str, float] = defaultdict(float)
    count_by_topic: dict[str, int] = defaultdict(int)
    examples: dict[str, dict] = {}
    for market in selected:
        volume = float(market.get("volume_fp") or 0)
        event_ticker = market.get("event_ticker") or ""
        category = event_categories.get(event_ticker, "")
        topic = classify_kalshi_market(market, category)
        volume_by_topic[topic] += volume
        count_by_topic[topic] += 1
        examples.setdefault(
            topic,
            {
                "title": market.get("title"),
                "volume": volume,
                "event_ticker": event_ticker,
                "category": category,
            },
        )

    total = sum(volume_by_topic.values())
    rows = []
    for topic in sorted(volume_by_topic, key=volume_by_topic.get, reverse=True):
        volume = volume_by_topic[topic]
        rows.append(
            {
                "date": as_of,
                "topic": topic,
                "volume1mo": volume,
                "share_pct": (volume / total * 100) if total else 0,
                "event_count": count_by_topic[topic],
                "example": examples.get(topic),
            }
        )

    return {
        "as_of": as_of,
        "platform": "Kalshi",
        "metric_label": "累计成交量",
        "source": KALSHI_MARKETS_URL,
        "params": {
            "limit": 500,
            "cursor": "paginated",
            "metric": "volume_fp",
        },
        "rows_fetched": len(markets),
        "positive_rows": len(positive_markets),
        "sample_rows": len(selected),
        "sum_volume1mo": total,
        "last_event_volume1mo": float(selected[-1].get("volume_fp") or 0) if selected else 0,
        "topics": rows,
    }


def write_current(summary: dict) -> None:
    write_platform_current(summary, "current", "current")


def write_platform_current(summary: dict, csv_stem: str, json_stem: str) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    current_csv = DATA_DIR / f"{csv_stem}.csv"
    current_json = DATA_DIR / f"{json_stem}.json"
    rows = current_rows_from_summary(summary, include_total=True)

    with current_csv.open("w", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(["date", "topic", "volume1mo", "share_pct", "event_count"])
        for row in rows:
            writer.writerow(
                [
                    row["date"],
                    row["topic"],
                    row["volume1mo"],
                    row["share_pct"],
                    row["event_count"],
                ]
            )

    current_payload = dict(summary)
    current_payload["topics_with_total"] = rows
    current_json.write_text(json.dumps(current_payload, indent=2), encoding="utf-8")


def update_history(summary: dict) -> list[dict]:
    return update_platform_history(summary, "history")


def update_platform_history(summary: dict, csv_stem: str) -> list[dict]:
    history_csv = DATA_DIR / f"{csv_stem}.csv"
    existing: list[dict] = []
    if history_csv.exists():
        with history_csv.open(newline="") as file:
            existing = list(csv.DictReader(file))

    as_of = summary["as_of"]
    existing = [row for row in existing if row.get("date") != as_of]
    for row in summary["topics"]:
        existing.append(
            {
                "date": row["date"],
                "topic": row["topic"],
                "volume1mo": f"{row['volume1mo']:.8f}",
                "share_pct": f"{row['share_pct']:.8f}",
                "event_count": str(row["event_count"]),
            }
        )

    existing.sort(key=lambda row: (row["date"], row["topic"]))
    with history_csv.open("w", newline="") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=["date", "topic", "volume1mo", "share_pct", "event_count"],
        )
        writer.writeheader()
        writer.writerows(existing)
    return existing


def current_rows_from_summary(summary: dict, include_total: bool = False) -> list[dict]:
    rows = [
        {
            "date": row["date"],
            "topic": row["topic"],
            "volume1mo": f"{row['volume1mo']:.8f}",
            "share_pct": f"{row['share_pct']:.8f}",
            "event_count": str(row["event_count"]),
        }
        for row in summary["topics"]
    ]
    if include_total:
        rows.append(
            {
                "date": summary["as_of"],
                "topic": "Total",
                "volume1mo": f"{summary['sum_volume1mo']:.8f}",
                "share_pct": "100.00000000",
                "event_count": str(sum(int(row["event_count"]) for row in rows)),
            }
        )
    return rows


def write_workbook(current_rows: list[dict], history_rows: list[dict]) -> None:
    write_platform_workbook(
        current_rows,
        history_rows,
        "polymarket_30d_topic_share.xlsx",
    )


def write_platform_workbook(
    current_rows: list[dict],
    history_rows: list[dict],
    workbook_name: str,
) -> None:
    workbook = Workbook()
    current_sheet = workbook.active
    current_sheet.title = "current"
    history_sheet = workbook.create_sheet("history")

    headers = ["date", "topic", "volume1mo", "share_pct", "event_count"]
    for sheet, rows in ((current_sheet, current_rows), (history_sheet, history_rows)):
        sheet.append(headers)
        for row in rows:
            sheet.append(
                [
                    row["date"],
                    row["topic"],
                    float(row["volume1mo"]),
                    float(row["share_pct"]),
                    int(row["event_count"]),
                ]
            )
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2D5A87")
            cell.alignment = Alignment(horizontal="center")
        for col_idx in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 18
        for row in sheet.iter_rows(min_row=2, min_col=3, max_col=4):
            row[0].number_format = '#,##0.00'
            row[1].number_format = '0.00'
        for row_idx in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_idx, column=2).value == "Total":
                for cell in sheet[row_idx]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="EAF2F8")

    workbook.save(DATA_DIR / workbook_name)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", default=date.today().isoformat())
    parser.add_argument("--limit", type=int, default=20)
    parser.add_argument("--max-rows", type=int, default=2000)
    parser.add_argument("--skip-kalshi", action="store_true")
    parser.add_argument("--kalshi-limit", type=int, default=500)
    parser.add_argument("--kalshi-max-pages", type=int, default=10)
    parser.add_argument("--kalshi-top-rows", type=int, default=2000)
    parser.add_argument(
        "--from-json",
        type=Path,
        help="Seed data from an existing summary JSON instead of fetching.",
    )
    args = parser.parse_args()

    as_of = args.date
    if args.from_json:
        summary = json.loads(args.from_json.read_text(encoding="utf-8"))
        summary["as_of"] = as_of
        for row in summary.get("categories", []):
            row["date"] = as_of
        for row in summary.get("topics", []):
            row["date"] = as_of
        if "topics" not in summary:
            summary = {
                "as_of": as_of,
                "source": summary.get("source", EVENTS_URL),
                "params": summary.get("params", {}),
                "rows_fetched": summary.get("rows_fetched", 0),
                "sum_volume1mo": summary.get("sum_volume1mo_top2000", 0),
                "last_event_volume1mo": summary.get("last_event_volume1mo", 0),
                "topics": [
                    {
                        "date": as_of,
                        "topic": row["topic"],
                        "volume1mo": row["volume1mo"],
                        "share_pct": row["share_pct"],
                        "event_count": row["event_count"],
                        "example": row.get("example"),
                    }
                    for row in summary.get("categories", [])
                ],
            }
    else:
        events = fetch_events(args.limit, args.max_rows)
        summary = summarize(events, as_of)

    write_current(summary)
    history_rows = update_history(summary)
    current_rows = current_rows_from_summary(summary, include_total=True)
    write_workbook(current_rows, history_rows)

    metadata = {
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "as_of": as_of,
        "rows_fetched": summary["rows_fetched"],
        "sum_volume1mo": summary["sum_volume1mo"],
        "last_event_volume1mo": summary["last_event_volume1mo"],
    }
    (DATA_DIR / "metadata.json").write_text(json.dumps(metadata, indent=2), encoding="utf-8")

    output = {"polymarket": metadata}

    if not args.skip_kalshi:
        kalshi_markets = fetch_kalshi_markets(args.kalshi_limit, args.kalshi_max_pages)
        kalshi_summary = summarize_kalshi(kalshi_markets, as_of, args.kalshi_top_rows)
        write_platform_current(kalshi_summary, "kalshi_current", "kalshi_current")
        kalshi_history_rows = update_platform_history(kalshi_summary, "kalshi_history")
        kalshi_current_rows = current_rows_from_summary(kalshi_summary, include_total=True)
        write_platform_workbook(
            kalshi_current_rows,
            kalshi_history_rows,
            "kalshi_cumulative_topic_share.xlsx",
        )
        kalshi_metadata = {
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "as_of": as_of,
            "rows_fetched": kalshi_summary["rows_fetched"],
            "positive_rows": kalshi_summary["positive_rows"],
            "sample_rows": kalshi_summary["sample_rows"],
            "sum_volume1mo": kalshi_summary["sum_volume1mo"],
            "last_event_volume1mo": kalshi_summary["last_event_volume1mo"],
            "metric_label": kalshi_summary["metric_label"],
        }
        (DATA_DIR / "kalshi_metadata.json").write_text(
            json.dumps(kalshi_metadata, indent=2),
            encoding="utf-8",
        )
        output["kalshi"] = kalshi_metadata

    print(json.dumps(output, indent=2))


if __name__ == "__main__":
    main()
